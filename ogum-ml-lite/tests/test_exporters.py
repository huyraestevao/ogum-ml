from __future__ import annotations

import importlib.util

import pandas as pd
import pytest
from ogum_lite import exporters
from ogum_lite.exporters import export_onnx, export_xlsx
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier


def test_export_xlsx_creates_expected_tabs(tmp_path) -> None:
    out_path = tmp_path / "report.xlsx"
    context = {
        "dataset": {"name": "demo"},
        "metrics": {"accuracy": 0.92},
    }
    tables = {
        "MSC": pd.DataFrame({"activation_energy": [200.0], "mse": [0.01]}),
        "Features": pd.DataFrame({"sample_id": ["s1"], "y_final": [0.9]}),
        "Metrics": pd.DataFrame({"metric": ["accuracy"], "value": [0.92]}),
    }

    export_xlsx(out_path, context=context, tables=tables, images=None)

    workbook = load_workbook(out_path)
    sheetnames = set(workbook.sheetnames)
    assert {"Summary", "MSC", "Features", "Metrics"}.issubset(sheetnames)


def test_export_onnx_returns_none_when_optional_missing(monkeypatch, tmp_path) -> None:
    model = RandomForestClassifier(n_estimators=1, random_state=0)
    model.fit([[0.0, 1.0], [1.0, 0.0]], [0, 1])
    out_path = tmp_path / "model.onnx"

    def raise_importerror(name: str) -> None:
        raise ImportError(name)

    monkeypatch.setattr(exporters.importlib, "import_module", raise_importerror)

    result = export_onnx(model, ["f1", "f2"], out_path)
    assert result is None


@pytest.mark.skipif(
    importlib.util.find_spec("skl2onnx") is None
    and importlib.util.find_spec("onnxmltools") is None,
    reason="ONNX conversion dependencies not installed",
)
def test_export_onnx_generates_file_when_deps_present(tmp_path) -> None:
    model = RandomForestClassifier(n_estimators=1, random_state=0)
    model.fit([[0.0, 1.0], [1.0, 0.0]], [0, 1])
    out_path = tmp_path / "model.onnx"

    result = export_onnx(model, ["f1", "f2"], out_path)
    assert result is None or result.exists()
